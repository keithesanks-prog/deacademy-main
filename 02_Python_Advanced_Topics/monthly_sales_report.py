from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
import random

def generate_sample_sales_data():
    """Generate sample sales data for the month."""
    products = ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Headphones', 'Webcam', 'USB Drive', 'Printer']
    regions = ['North', 'South', 'East', 'West']
    
    sales_data = []
    start_date = datetime(2026, 1, 1)
    
    for day in range(31):
        current_date = start_date + timedelta(days=day)
        for _ in range(random.randint(3, 8)):  # 3-8 sales per day
            sale = {
                'Date': current_date.strftime('%Y-%m-%d'),
                'Product': random.choice(products),
                'Region': random.choice(regions),
                'Quantity': random.randint(1, 10),
                'Unit Price': random.randint(20, 500),
            }
            sale['Total'] = sale['Quantity'] * sale['Unit Price']
            sales_data.append(sale)
    
    return sales_data

def create_monthly_sales_report():
    """Create a formatted monthly sales report in Excel."""
    
    # Generate sample data
    sales_data = generate_sample_sales_data()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Define styles
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    title_font = Font(name='Calibri', size=18, bold=True, color='366092')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'MONTHLY SALES REPORT - JANUARY 2026'
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    # Add report generation date
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)
    
    # Add headers
    headers = ['Date', 'Product', 'Region', 'Quantity', 'Unit Price', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Add data
    for row_idx, sale in enumerate(sales_data, start=5):
        ws.cell(row=row_idx, column=1, value=sale['Date'])
        ws.cell(row=row_idx, column=2, value=sale['Product'])
        ws.cell(row=row_idx, column=3, value=sale['Region'])
        ws.cell(row=row_idx, column=4, value=sale['Quantity'])
        ws.cell(row=row_idx, column=5, value=sale['Unit Price'])
        ws.cell(row=row_idx, column=6, value=sale['Total'])
        
        # Apply borders to data cells
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = border
            
        # Format currency columns
        ws.cell(row=row_idx, column=5).number_format = '$#,##0'
        ws.cell(row=row_idx, column=6).number_format = '$#,##0'
        
        # Center align quantity
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
    
    # Add summary section
    last_row = len(sales_data) + 5
    summary_row = last_row + 2
    
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    ws[f'A{summary_row}'] = 'SUMMARY'
    ws[f'A{summary_row}'].font = Font(bold=True, size=12, color='366092')
    
    # Calculate totals
    total_sales = sum(sale['Total'] for sale in sales_data)
    total_quantity = sum(sale['Quantity'] for sale in sales_data)
    
    summary_data = [
        ('Total Sales:', f'${total_sales:,}'),
        ('Total Units Sold:', total_quantity),
        ('Number of Transactions:', len(sales_data)),
        ('Average Transaction:', f'${total_sales/len(sales_data):.2f}'),
    ]
    
    for idx, (label, value) in enumerate(summary_data, start=summary_row+1):
        ws[f'A{idx}'] = label
        ws[f'A{idx}'].font = Font(bold=True)
        ws[f'B{idx}'] = value
        ws[f'B{idx}'].font = Font(size=11)
    
    # Create Summary by Product sheet
    ws_product = wb.create_sheet("By Product")
    ws_product.merge_cells('A1:C1')
    ws_product['A1'] = 'SALES BY PRODUCT'
    ws_product['A1'].font = title_font
    ws_product['A1'].alignment = title_alignment
    
    # Headers for product summary
    product_headers = ['Product', 'Units Sold', 'Total Revenue']
    for col, header in enumerate(product_headers, start=1):
        cell = ws_product.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Calculate product totals
    product_totals = {}
    for sale in sales_data:
        product = sale['Product']
        if product not in product_totals:
            product_totals[product] = {'quantity': 0, 'revenue': 0}
        product_totals[product]['quantity'] += sale['Quantity']
        product_totals[product]['revenue'] += sale['Total']
    
    # Add product data
    row_idx = 4
    for product, totals in sorted(product_totals.items(), key=lambda x: x[1]['revenue'], reverse=True):
        ws_product.cell(row=row_idx, column=1, value=product)
        ws_product.cell(row=row_idx, column=2, value=totals['quantity'])
        ws_product.cell(row=row_idx, column=3, value=totals['revenue'])
        ws_product.cell(row=row_idx, column=3).number_format = '$#,##0'
        
        for col in range(1, 4):
            ws_product.cell(row=row_idx, column=col).border = border
            ws_product.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
        
        row_idx += 1
    
    # Add chart
    chart = BarChart()
    chart.title = "Revenue by Product"
    chart.style = 10
    chart.y_axis.title = 'Revenue ($)'
    chart.x_axis.title = 'Product'
    
    data = Reference(ws_product, min_col=3, min_row=3, max_row=row_idx-1)
    cats = Reference(ws_product, min_col=1, min_row=4, max_row=row_idx-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws_product.add_chart(chart, "E3")
    
    # Create Summary by Region sheet
    ws_region = wb.create_sheet("By Region")
    ws_region.merge_cells('A1:C1')
    ws_region['A1'] = 'SALES BY REGION'
    ws_region['A1'].font = title_font
    ws_region['A1'].alignment = title_alignment
    
    # Headers for region summary
    region_headers = ['Region', 'Units Sold', 'Total Revenue']
    for col, header in enumerate(region_headers, start=1):
        cell = ws_region.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Calculate region totals
    region_totals = {}
    for sale in sales_data:
        region = sale['Region']
        if region not in region_totals:
            region_totals[region] = {'quantity': 0, 'revenue': 0}
        region_totals[region]['quantity'] += sale['Quantity']
        region_totals[region]['revenue'] += sale['Total']
    
    # Add region data
    row_idx = 4
    for region, totals in sorted(region_totals.items(), key=lambda x: x[1]['revenue'], reverse=True):
        ws_region.cell(row=row_idx, column=1, value=region)
        ws_region.cell(row=row_idx, column=2, value=totals['quantity'])
        ws_region.cell(row=row_idx, column=3, value=totals['revenue'])
        ws_region.cell(row=row_idx, column=3).number_format = '$#,##0'
        
        for col in range(1, 4):
            ws_region.cell(row=row_idx, column=col).border = border
            ws_region.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
        
        row_idx += 1
    
    # Adjust column widths for all sheets
    from openpyxl.utils import get_column_letter
    
    for sheet in [ws, ws_product, ws_region]:
        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and not isinstance(cell, type(sheet['A1'])):  # Skip merged cells
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 12
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    filename = f'Monthly_Sales_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    print(f"✓ Report generated successfully: {filename}")
    print(f"✓ Total Sales: ${total_sales:,}")
    print(f"✓ Total Transactions: {len(sales_data)}")
    print(f"✓ Sheets created: Sales Report, By Product, By Region")
    
    return filename

if __name__ == "__main__":
    print("=" * 60)
    print("MONTHLY SALES REPORT GENERATOR")
    print("=" * 60)
    print("\nGenerating report...")
    create_monthly_sales_report()
    print("\n" + "=" * 60)
